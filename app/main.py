# main.py - Fixed with BackgroundTasks (No Celery Needed)
# ============================================================================

import os
import asyncio
import logging
import traceback
from contextlib import asynccontextmanager
from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, status, Request, BackgroundTasks
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, create_engine
from sqlalchemy.orm import sessionmaker
from typing import List, Optional
from pathlib import Path
from datetime import datetime
from decimal import Decimal
import re

from app.core.config import settings
from app.core.database import get_db, init_db
from app.models.user import User
from app.models.invoice import Invoice, InvoiceStatus, LineItem
from app.services.auth import AuthService
from app.services.storage import StorageService
from app.services.payment import PaymentService
from app.schemas.auth import LoginRequest, TokenResponse
from app.schemas.invoice import InvoiceResponse, InvoiceUploadResponse

# Setup logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ============================================================================
# BACKGROUND PROCESSING FUNCTION
# ============================================================================

def process_invoice_background(invoice_id: int):
    """Process invoice in background thread (synchronous)"""
    
    # Create synchronous database session
    sync_db_url = settings.DATABASE_URL.replace('postgresql+asyncpg://', 'postgresql://')
    engine = create_engine(sync_db_url)
    SessionLocal = sessionmaker(bind=engine)
    db = SessionLocal()
    
    try:
        logger.info(f"🔄 Processing invoice {invoice_id}")
        
        # Get invoice
        invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
        if not invoice:
            logger.error(f"Invoice {invoice_id} not found")
            return
        
        # Update to processing
        invoice.status = InvoiceStatus.PROCESSING
        db.commit()
        
        # Import libraries
        import fitz  # PyMuPDF
        import pdfplumber
        import cv2
        import numpy as np
        import pytesseract
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        file_path = invoice.original_path
        file_ext = Path(file_path).suffix.lower()
        
        # Initialize result
        extracted_data = {
            'invoice_number': None,
            'vendor_name': None,
            'invoice_date': None,
            'total_amount': None,
            'line_items': []
        }
        
        full_text = ""
        
        # PROCESS FILE
        if file_ext == '.pdf':
            logger.info("Processing PDF...")
            try:
                with pdfplumber.open(file_path) as pdf:
                    for page in pdf.pages:
                        page_text = page.extract_text()
                        if page_text:
                            full_text += page_text + "\n"
            except Exception as e:
                logger.warning(f"pdfplumber failed: {e}, using OCR...")
                doc = fitz.open(file_path)
                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    pix = page.get_pixmap(dpi=300)
                    img_data = pix.tobytes("png")
                    img_array = np.frombuffer(img_data, np.uint8)
                    img = cv2.imdecode(img_array, cv2.IMREAD_COLOR)
                    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
                    text = pytesseract.image_to_string(gray, config='--oem 3 --psm 6')
                    full_text += text + "\n"
                doc.close()
        
        elif file_ext in ['.jpg', '.jpeg', '.png']:
            logger.info("Processing image...")
            img = cv2.imread(file_path)
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            clahe = cv2.createCLAHE(clipLimit=2.0, tileGridSize=(8,8))
            gray = clahe.apply(gray)
            binary = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
            full_text = pytesseract.image_to_string(binary, config='--oem 3 --psm 6')
        
        logger.info(f"Extracted {len(full_text)} characters")
        
        # EXTRACT INVOICE NUMBER
        patterns = [
            r'invoice\s*#?\s*:?\s*([A-Z0-9\-]+)',
            r'inv\s*#?\s*:?\s*([A-Z0-9\-]+)',
            r'#\s*([A-Z0-9]{3,})',
        ]
        for pattern in patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                extracted_data['invoice_number'] = match.group(1)
                break
        
        # EXTRACT DATE
        date_patterns = [
            r'date\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})',
            r'(\d{1,2}[/-]\d{1,2}[/-]\d{4})',
            r'(\d{4}-\d{2}-\d{2})',
        ]
        for pattern in date_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                date_str = match.group(1)
                for fmt in ['%m/%d/%Y', '%d/%m/%Y', '%Y-%m-%d', '%m-%d-%Y']:
                    try:
                        extracted_data['invoice_date'] = datetime.strptime(date_str, fmt)
                        break
                    except:
                        continue
                if extracted_data['invoice_date']:
                    break
        
        # EXTRACT VENDOR
        lines = [line.strip() for line in full_text.split('\n') if line.strip()]
        for line in lines[:8]:
            if re.search(r'invoice|bill|receipt|date|total', line, re.IGNORECASE):
                continue
            if len(line) > 3 and len(line) < 100:
                num_ratio = sum(c.isdigit() for c in line) / len(line) if line else 0
                if num_ratio < 0.5:
                    extracted_data['vendor_name'] = line
                    break
        
        # EXTRACT TOTAL
        amount_patterns = [
            r'total\s*:?\s*\$?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',
            r'amount\s*due\s*:?\s*\$?\s*(\d+(?:,\d{3})*(?:\.\d{2})?)',
        ]
        amounts = []
        for pattern in amount_patterns:
            matches = re.findall(pattern, full_text, re.IGNORECASE)
            for amt in matches:
                try:
                    amounts.append(float(amt.replace(',', '')))
                except:
                    pass
        if amounts:
            extracted_data['total_amount'] = Decimal(str(max(amounts)))
        
        # EXTRACT LINE ITEMS
        for line in lines:
            amount_match = re.search(r'\$?\s*(\d+(?:,\d{3})*\.\d{2})\s*$', line)
            if amount_match:
                amount = amount_match.group(1).replace(',', '')
                desc = line[:amount_match.start()].strip()
                if desc and len(desc) > 3:
                    extracted_data['line_items'].append({
                        'description': desc,
                        'quantity': None,
                        'unit_price': None,
                        'total_amount': Decimal(amount)
                    })
        
        logger.info(f"Extracted: {extracted_data['invoice_number']}, {extracted_data['vendor_name']}, ${extracted_data['total_amount']}")
        
        # UPDATE DATABASE
        invoice.invoice_number = extracted_data.get('invoice_number')
        invoice.vendor_name = extracted_data.get('vendor_name')
        invoice.invoice_date = extracted_data.get('invoice_date')
        invoice.total_amount = extracted_data.get('total_amount')
        
        # Clear old line items
        db.query(LineItem).filter(LineItem.invoice_id == invoice.id).delete()
        
        # Add new line items
        for item in extracted_data['line_items'][:50]:
            line_item = LineItem(
                invoice_id=invoice.id,
                description=item.get('description', 'N/A'),
                quantity=item.get('quantity'),
                unit_price=item.get('unit_price'),
                tax_amount=item.get('tax_amount'),
                total_amount=item.get('total_amount', Decimal('0.00'))
            )
            db.add(line_item)
        
        db.commit()
        
        # GENERATE EXCEL
        from app.services.storage import StorageService
        storage = StorageService()
        excel_path = storage.get_excel_path(invoice.id, invoice.filename)
        
        # Create directory if needed
        Path(excel_path).parent.mkdir(parents=True, exist_ok=True)
        
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Invoice Data"
        
        # Header
        ws['A1'] = "Invoice Information"
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = "Invoice Number:"
        ws[f'B{row}'] = extracted_data.get('invoice_number', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Vendor:"
        ws[f'B{row}'] = extracted_data.get('vendor_name', 'N/A')
        row += 1
        
        ws[f'A{row}'] = "Date:"
        ws[f'B{row}'] = extracted_data['invoice_date'].strftime('%Y-%m-%d') if extracted_data.get('invoice_date') else 'N/A'
        row += 1
        
        ws[f'A{row}'] = "Total:"
        ws[f'B{row}'] = f"${extracted_data.get('total_amount', 0)}"
        row += 2
        
        # Line items
        if extracted_data['line_items']:
            ws[f'A{row}'] = "Line Items"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            headers = ['Description', 'Quantity', 'Unit Price', 'Total']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E8F5E8", end_color="E8F5E8", fill_type="solid")
            row += 1
            
            for item in extracted_data['line_items']:
                ws.cell(row=row, column=1, value=item.get('description', ''))
                ws.cell(row=row, column=2, value=str(item.get('quantity', '')))
                ws.cell(row=row, column=3, value=f"${item.get('unit_price', '')}" if item.get('unit_price') else '')
                ws.cell(row=row, column=4, value=f"${item.get('total_amount', '')}")
                row += 1
        
        # Auto-adjust columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        workbook.save(excel_path)
        workbook.close()
        
        invoice.excel_path = excel_path
        invoice.status = InvoiceStatus.COMPLETED
        invoice.error_message = None
        db.commit()
        
        logger.info(f"✅ Invoice {invoice_id} completed!")
        
    except Exception as e:
        logger.error(f"❌ Error processing invoice {invoice_id}: {e}")
        logger.error(traceback.format_exc())
        try:
            invoice = db.query(Invoice).filter(Invoice.id == invoice_id).first()
            if invoice:
                invoice.status = InvoiceStatus.FAILED
                invoice.error_message = str(e)[:500]
                db.commit()
        except:
            pass
    
    finally:
        db.close()

# ============================================================================
# FASTAPI APP
# ============================================================================

@asynccontextmanager
async def lifespan(app: FastAPI):
    try:
        await init_db()
        logger.info("✅ Database initialized")
    except Exception as e:
        logger.error(f"❌ Database init failed: {e}")
    yield

app = FastAPI(
    title="InvoiceAI API",
    description="AI-powered invoice processing",
    version="1.0.0",
    lifespan=lifespan
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.middleware("http")
async def add_coop_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers["Cross-Origin-Opener-Policy"] = "same-origin-allow-popups"
    return response

security = HTTPBearer()

async def get_current_user(
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: AsyncSession = Depends(get_db)
) -> User:
    auth_service = AuthService()
    user = await auth_service.get_current_user(credentials.credentials, db)
    if not user:
        raise HTTPException(status_code=401, detail="Invalid auth")
    return user

# ============================================================================
# ENDPOINTS
# ============================================================================

@app.get("/health")
async def health():
    return {"status": "healthy"}

@app.post("/api/auth/login")
async def login(request: LoginRequest, db: AsyncSession = Depends(get_db)):
    try:
        auth_service = AuthService()
        if request.provider == "google":
            if not request.token:
                raise HTTPException(status_code=400, detail="Token required")
            return await auth_service.google_login(request.token, db)
        elif request.provider == "magic_link":
            if not request.email:
                raise HTTPException(status_code=400, detail="Email required")
            return await auth_service.send_magic_link(request.email, db)
        else:
            raise HTTPException(status_code=400, detail="Invalid provider")
    except Exception as e:
        logger.error(f"Login error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/auth/verify")
async def verify_magic_link(token: str, db: AsyncSession = Depends(get_db)):
    try:
        auth_service = AuthService()
        return await auth_service.verify_magic_link(token, db)
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/invoices/upload", response_model=List[InvoiceUploadResponse])
async def upload_invoices(
    background_tasks: BackgroundTasks,
    files: List[UploadFile] = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    """Upload invoices - uses BackgroundTasks for processing"""
    if len(files) > 100:
        raise HTTPException(status_code=400, detail="Max 100 files")
    
    if current_user.credits_balance < len(files) and current_user.plan == "credit_pack":
        raise HTTPException(status_code=402, detail="Insufficient credits")
    
    storage_service = StorageService()
    responses = []
    
    for file in files:
        if file.content_type not in ["application/pdf", "image/jpeg", "image/png"]:
            continue
        
        file_path = await storage_service.save_upload(file, current_user.id)
        
        invoice = Invoice(
            user_id=current_user.id,
            filename=file.filename,
            original_path=file_path,
            status=InvoiceStatus.PENDING
        )
        db.add(invoice)
        await db.commit()
        await db.refresh(invoice)
        
        # Add to background tasks
        background_tasks.add_task(process_invoice_background, invoice.id)
        
        responses.append(InvoiceUploadResponse(
            invoice_id=invoice.id,
            filename=file.filename,
            status="queued"
        ))
    
    return responses

@app.get("/api/invoices", response_model=List[InvoiceResponse])
async def list_invoices(
    skip: int = 0,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(
        select(Invoice)
        .where(Invoice.user_id == current_user.id)
        .offset(skip)
        .limit(limit)
        .order_by(Invoice.created_at.desc())
    )
    invoices = result.scalars().all()
    return [InvoiceResponse.model_validate(inv) for inv in invoices]

@app.get("/api/invoices/status/{invoice_id}")
async def get_status(
    invoice_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    invoice = await db.get(Invoice, invoice_id)
    if not invoice or invoice.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Not found")
    return InvoiceResponse.model_validate(invoice)

@app.get("/api/invoices/download/{invoice_id}")
async def download_excel(
    invoice_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db)
):
    invoice = await db.get(Invoice, invoice_id)
    if not invoice or invoice.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Not found")
    
    if invoice.status != InvoiceStatus.COMPLETED or not invoice.excel_path:
        raise HTTPException(status_code=400, detail="Not ready")
    
    return FileResponse(
        invoice.excel_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{invoice.filename.rsplit('.', 1)[0]}.xlsx"
    )

@app.get("/")
async def root():
    return {"message": "InvoiceAI API", "version": "1.0.0"}
